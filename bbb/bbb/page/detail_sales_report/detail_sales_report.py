# Copyright (c) 2022, invento software limited and contributors
# For license information, please see license.txt

import frappe
import os
import io
import requests
from frappe import _
import json
from datetime import datetime
import time
from frappe.utils import get_time
import openpyxl
from openpyxl import Workbook
from frappe.utils.file_manager import get_file_path
from six.moves.urllib.parse import urljoin
from openpyxl.styles import Alignment, PatternFill, Side, Border, Font
from openpyxl.utils import get_column_letter
from frappe.utils.file_manager import get_file_path


def get_conditions(filters):
    conditions = []

    if filters.get("from_date"):
        conditions.append("sales_invoice.posting_date >= '%s'" % filters.get("from_date"))

    if filters.get("to_date"):
        conditions.append("sales_invoice.posting_date <= '%s'" % filters.get("to_date"))

    if filters.get("outlet"):
        conditions.append("sales_invoice.pos_profile = '%s'" % filters.get("outlet"))

    if conditions:
        conditions = " and ".join(conditions)
    return conditions


@frappe.whitelist()
def get_invoice_data(filters):
    try:
        filters = json.loads(filters)
    except:
        pass

    conditions = get_conditions(filters)
    invoice_type = filters.get('switch_invoice', "Sales Invoice")
    query_result = frappe.db.sql("""
    		select
    			sales_invoice.pos_profile, sales_invoice.total_taxes_and_charges as vat, sales_invoice.name, sales_invoice.posting_date, sales_invoice.posting_time, 
    			sales_invoice_item.price_list_rate as unit_price, sales_invoice_item.rate as selling_rate, item.standard_rate as mrp, sales_invoice_item.idx,
    			sales_invoice_item.qty as quantity, sales_invoice_item.item_name, item.standard_rate as mrp, sales_invoice.served_by,
    			(sales_invoice_item.qty * item.standard_rate) as mrp_total, ((sales_invoice_item.qty * sales_invoice_item.rate) - (sales_invoice_item.qty * item.buying_rate)) as profit_loss, 
    			(sales_invoice_item.qty * item.buying_rate) as buying_total, sales_invoice.customer_name, sales_invoice.customer_mobile_number,
    			((sales_invoice_item.qty * sales_invoice_item.discount_amount)) as discount, sales_invoice.set_warehouse, sales_invoice.special_note, 
    			(sales_invoice.total - sales_invoice.net_total) as special_discount, sales_invoice.total_qty, sales_invoice.return_against,
    			 sales_invoice_item.net_amount, sales_invoice_item.amount as total_amount, sales_invoice.customer_name, sales_invoice.is_return,
    			 sales_invoice.total, sales_invoice.rounded_total, sales_invoice.total_taxes_and_charges, sales_invoice.net_total, sales_invoice.rounding_adjustment
    		from `tab%s` sales_invoice, `tab%s Item` sales_invoice_item, `tabItem` item
    		where sales_invoice.name = sales_invoice_item.parent and item.item_code = sales_invoice_item.item_code
    			and sales_invoice.docstatus = 1 and %s
    		order by sales_invoice.posting_date, sales_invoice.posting_time asc 
    		""" % (invoice_type, invoice_type, conditions), as_dict=1)

    data = {}
    for index, result in enumerate(query_result):
        if data.get(result.get('name'), None):
            sales_data = data.get(result.get('name'))
            if result['is_return'] == 0:
                sales_data['mrp_total_price'] = sales_data['mrp_total_price'] + result['mrp_total']
                sales_data['total_selling_rate'] = sales_data['total_selling_rate'] + result['selling_rate']
                sales_data['total_amount_'] = sales_data['total_amount_'] + result['total_amount']
                sales_data['total_discount'] = sales_data['total_discount'] + result['discount']
            else:
                sales_data['mrp_total'] = ''
                sales_data['total_selling_rate'] = ''
                sales_data['total_mrp_price'] = ''
                sales_data['total_amount_'] = ''
                sales_data['total_discount'] = ''

            sales_data['profit_loss'] = sales_data['profit_loss'] + result['profit_loss']
            sales_data['total_item_row'] = sales_data['total_item_row'] + 1

        elif data.get(result['return_against'], None):


            sales_data = data.get(result['return_against'])

            if sales_data['exchange_item'] < 0:
                sales_data['total_item_row'] = sales_data['total_item_row'] + 1
                sales_data['exchange_item'] = result['idx']
                if result['special_note']:
                    sales_data['total_item_row'] = sales_data['total_item_row'] + 1

            sales_data['total_item_row'] = sales_data['total_item_row'] + 1
            sales_data['total_return_amount'] = result['rounded_total']


        else:
            result['exchange_item'] = -1


            if result['is_return'] < 0:
                result['exchange_item'] = result['idx']

            if result['special_note']:
                result['total_item_row'] = 4
            else:
                result['total_item_row'] = 3

            if result['is_return'] == 0:
                result['total_selling_rate'] = result['selling_rate']
                result['total_discount'] = result['discount']
                result['total_amount_'] = result['total_amount']
                result['total_mrp_price'] = result['mrp']
                result['total_return_amount'] = ''
            else:
                result['total_selling_rate'] = ''
                result['total_discount'] = ''
                result['total_amount_'] = ''
                result['total_mrp_price'] = ''
                result['total_return_amount'] = result['rounded_total']
            result['mrp_total_price'] = result['mrp_total']
            data[result.get('name')] = result

    return data, query_result


@frappe.whitelist()
def generate_table_data(filters):
    data, query_result = get_invoice_data(filters)
    sales_data = {}
    total_item_qty = 0
    total_exchange_qty = 0
    unit_total_amount = 0
    product_discount_amount = 0
    rounded_total_amount = 0
    special_disc_amount = 0
    total_exchange_amount = 0

    for index, result in enumerate(query_result):
        sl_number = len(sales_data) + 1
        next_invoice = True
        invoice_name = result.get('name')

        try:
            next_invoice = query_result[index + 1]
            next_invoice = query_result[index + 1]['name']
        except:
            next_invoice = not next_invoice

        if result.get('is_return') == 1:
            total_exchange_qty += result.get('quantity')
        else:
            total_item_qty += result.get('quantity')

        unit_total_amount += result.get('mrp') * result.get('quantity')
        product_discount_amount += result.get('discount')
        invoice_item_total = data.get(invoice_name)
        if invoice_item_total is None:
            invoice_item_total = data.get(result['return_against'])

        time_obj = datetime.strptime(f"{result['posting_time']}", '%H:%M:%S.%f').strftime("%I:%M %p")
        style = 'style="background:#ffff0085"' if result['special_discount'] > 0 else ''
        exchange_amount = 0 if invoice_item_total['total_return_amount'] == '' else invoice_item_total['total_return_amount']
        total = result['rounded_total'] + exchange_amount


        tr_first_row = f"""
            <tr>
                <th scope='row' rowspan='{invoice_item_total['total_item_row']}'>{sl_number}</th>
                <td rowspan='{invoice_item_total['total_item_row']}' class="text-center">{invoice_name}<br>{result['customer_name']} ({result['customer_mobile_number']})<br>{result['posting_date']} {time_obj}</td>
                <td>Name</td>
                <td>MRP</td>
                <td>Qty</td>
                <td>MRP Total</td>
                <td>Disc</td>
                <td>Total</td>
                <td rowspan='{invoice_item_total['total_item_row']}'>{result['served_by']}</td>
                <td rowspan='{invoice_item_total['total_item_row']}'>{result['set_warehouse']}</td>
                <td {style} rowspan='{invoice_item_total['total_item_row']}'>{result['special_discount']}</td>
                <td rowspan='{invoice_item_total['total_item_row']}'>{result['rounding_adjustment']}</td>
                <td rowspan='{invoice_item_total['total_item_row']}'>{result['total_taxes_and_charges']}</td>
                <td rowspan='{invoice_item_total['total_item_row']}'>{result['rounded_total']}</td>
                <td rowspan='{invoice_item_total['total_item_row']}'>{invoice_item_total['total_return_amount']}</td>
                <td rowspan='{invoice_item_total['total_item_row']}'>{total}</td>
            </tr>

            <tr>
                <td>{result.get('item_name')}</td>
                <td>{result.get('mrp')}</td>
                <td>{result.get('quantity')}</td>
                <td>{result.get('mrp_total')}</td>
                <td>{result.get('discount')}</td>
                <td>{result.get('total_amount')}</td>
            </tr>
        """

        tr_return_with_header = f"""
            <tr>
                <th scope='row' rowspan='{invoice_item_total['total_item_row']}'>{sl_number}</th>
                <td rowspan='{invoice_item_total['total_item_row']}' class="text-center">{invoice_name}<br>{result['customer_name']} ({result['customer_mobile_number']})<br>{result['posting_date']} {time_obj}</td>
                <td>Name</td>
                <td>MRP</td>
                <td>Qty</td>
                <td>MRP Total</td>
                <td>Disc</td>
                <td>Total</td>
                <td rowspan='{invoice_item_total['total_item_row']}'>{result['served_by']}</td>
                <td rowspan='{invoice_item_total['total_item_row']}'>{result['set_warehouse']}</td>
                <td {style} rowspan='{invoice_item_total['total_item_row']}'>{result['special_discount']}</td>
                <td rowspan='{invoice_item_total['total_item_row']}'>{invoice_item_total['rounding_adjustment']}</td>
                <td rowspan='{invoice_item_total['total_item_row']}'>{result['total_taxes_and_charges']}</td>
                <td rowspan='{invoice_item_total['total_item_row']}'>{result['rounded_total']}</td>
                <td rowspan='{invoice_item_total['total_item_row']}'>{invoice_item_total['total_return_amount']}</td>
                <td rowspan='{invoice_item_total['total_item_row']}'>{total}</td>
            </tr>
            <tr>
                <td colspan='6' class="font-weight-bold text-center" style="background:#c3e0aad4">Exchange Product List</td>
            </tr>
            <tr>
                <td>{result.get('item_name')}</td>
                <td style="background:#ffff0085">{result.get('mrp')}</td>
                <td style="background:#ffff0085">{result.get('quantity')}</td>
                <td style="background:#ffff0085">{result.get('selling_rate')}</td>
                <td style="background:#ffff0085">{result.get('discount')}</td>
                <td style="background:#ffff0085">{result.get('total_amount')}</td>
            </tr>
        """

        tr_rowspan = f"""
            <tr>
                <td>{result.get('item_name')}</td>
                <td>{result.get('mrp')}</td>
                <td>{result.get('quantity')}</td>
                <td>{result.get('mrp_total')}</td>
                <td>{result.get('discount')}</td>
                <td>{result.get('total_amount')}</td>
            </tr>
        """

        tr_total_row = f"""
            <tr>
                <td colspan='2'>Total</td>     
                <td>{invoice_item_total.get('total_qty')}</td>     
                <td>{invoice_item_total.get('mrp_total_price')}</td>     
                <td>{invoice_item_total.get('total_discount')}</td>     
                <td>{invoice_item_total.get('total_amount_')}</td>   
            </tr>  
        """

        tr_special_note = f"""
            <tr style="background:#c3e0aad4">
                <td colspan='6'>Note: {result['special_note']}</td>  
            </tr>
        """
        tr_return_head = """
            <tr>
                <td colspan='6' class="font-weight-bold text-center" style="background:#c3e0aad4">Exchange Product List</td>
            </tr>
        """
        tr_return_row = f"""
            <tr>
                <td>{result.get('item_name')}</td>
                <td style="background:#ffff0085">{result.get('mrp')}</td>
                <td style="background:#ffff0085">{result.get('quantity')}</td>
                <td style="background:#ffff0085">{result.get('mrp_total')}</td>
                <td style="background:#ffff0085">{result.get('discount')}</td>
                <td style="background:#ffff0085">{result.get('total_amount')}</td>
            </tr>
        """

        if next_invoice and invoice_name == next_invoice:
            if sales_data.get(invoice_name):
                if result['is_return'] == 1:
                    sales_data[invoice_name].append(tr_return_row)
                else:
                    sales_data[invoice_name].append(tr_rowspan)
            else:
                if result['is_return'] == 1:
                    if sales_data == {}:
                        sales_data[invoice_name] = [tr_return_with_header]
                    elif sales_data.get(result['return_against']):
                        if invoice_item_total['exchange_item'] == result['idx']:
                            sales_data[result['return_against']].append(tr_return_head)
                        sales_data[result['return_against']].append(tr_return_row)
                    else:
                        sales_data[invoice_name] = [tr_return_with_header]
                else:
                    sales_data[invoice_name] = [tr_first_row]
                    rounded_total_amount += result['rounded_total']
                    total_exchange_amount += 0 if invoice_item_total['total_return_amount'] == '' else invoice_item_total['total_return_amount']
                    special_disc_amount += result['special_discount']

        elif next_invoice and invoice_name != next_invoice:
            if sales_data.get(invoice_name):
                if result['is_return'] == 1:
                    sales_data[invoice_name].append(tr_return_row)
                else:
                    sales_data[invoice_name].append(tr_rowspan)
                if result['is_return'] == 0:
                    sales_data[invoice_name].append(tr_total_row)
                if result['special_note']:
                    sales_data[invoice_name].append(tr_special_note)
            else:
                if result['is_return'] == 1:
                    if sales_data == {}:
                        if invoice_item_total['exchange_item'] == result['idx']:
                            sales_data[invoice_name].append(tr_return_head)
                        sales_data[invoice_name] = [tr_return_with_header]
                        if result['special_note']:
                            sales_data[invoice_name].append(tr_special_note)
                    elif sales_data.get(result['return_against']):
                        if invoice_item_total['exchange_item'] == result['idx']:
                            sales_data[result['return_against']].append(tr_return_head)
                        sales_data[result['return_against']].append(tr_return_row)
                        if result['special_note']:
                            sales_data[result['return_against']].append(tr_special_note)
                    else:
                        sales_data[invoice_name] = [tr_return_with_header]
                    # if result['special_note']:
                    #     sales_data[result['return_against']].append(tr_special_note)
                else:
                    sales_data[invoice_name] = [tr_first_row]
                    rounded_total_amount += result['rounded_total']
                    total_exchange_amount += 0 if invoice_item_total['total_return_amount'] == '' else invoice_item_total['total_return_amount']
                    special_disc_amount += result['special_discount']
                    sales_data[invoice_name].append(tr_total_row)
                    if result['special_note']:
                        sales_data[invoice_name].append(tr_special_note)

        elif next_invoice == False:
            if sales_data.get(invoice_name):
                if result['is_return'] == 1:
                    sales_data[invoice_name].append(tr_return_row)
                else:
                    sales_data[invoice_name].append(tr_rowspan)
                if result['is_return'] == 0:
                    sales_data[invoice_name].append(tr_total_row)
                if result['special_note']:
                    sales_data[invoice_name].append(tr_special_note)

            else:
                if result['is_return'] == 1:
                    if sales_data == {}:
                        if invoice_item_total['exchange_item'] == result['idx']:
                            sales_data[invoice_name].append(tr_return_head)
                        sales_data[invoice_name] = [tr_return_with_header]
                        if result['special_note']:
                            sales_data[invoice_name].append(tr_special_note)
                    elif sales_data.get(result['return_against']):
                        if invoice_item_total['exchange_item'] == result['idx']:
                            sales_data[result['return_against']].append(tr_return_head)
                        sales_data[result['return_against']].append(tr_return_row)
                        if result['special_note']:
                            sales_data[result['return_against']].append(tr_special_note)
                    else:
                        sales_data[invoice_name] = [tr_return_with_header]
                    # sales_data[result['return_against']].append(tr_special_note)
                else:
                    sales_data[invoice_name] = [tr_first_row]
                    rounded_total_amount += result['rounded_total']
                    total_exchange_amount += 0 if invoice_item_total['total_return_amount'] == '' else invoice_item_total['total_return_amount']
                    special_disc_amount += result['special_discount']
                    sales_data[invoice_name].append(tr_total_row)
                    if result['special_note']:
                        sales_data[result['return_against']].append(tr_special_note)
    sales_data['total_amount_row'] = [f"""
                        <tr class='font-weight-bold'>
                            <td rowspan='2'>Total </td>
                            <td rowspan='2'>Total Qty={total_item_qty} <br> Total Exchange Qty={total_exchange_qty}</td>
                            <td rowspan='2' colspan='6'>Unit Total={unit_total_amount} <br>Product Disc={product_discount_amount}</td>
                            <td rowspan='2' >Invoice Total={unit_total_amount - product_discount_amount}</td>
                            <td rowspan='2' >Special Discount={special_disc_amount}</td>
                            <td rowspan='2' colspan='2'>Sale Amount={rounded_total_amount}</td>
                            <td rowspan='2' colspan='2'>Exchange Amount={total_exchange_amount}</td>
                            <td rowspan='2' colspan='2'>Actual Sale Amount={rounded_total_amount + total_exchange_amount}</td>
                        </tr>
                        """]
    return sales_data

font = Font(size=9)
def get_item_description_header(ws, row_count):
    ws.cell(row=row_count, column=3).value = 'Name'
    ws.cell(row=row_count, column=4).value = 'MRP'
    ws.cell(row=row_count, column=5).value = 'QTY'
    ws.cell(row=row_count, column=6).value = 'MRP Total'
    ws.cell(row=row_count, column=7).value = 'Disc'
    ws.cell(row=row_count, column=8).value = 'Total'

    ws.cell(row=row_count, column=3).font = font
    ws.cell(row=row_count, column=4).font = font
    ws.cell(row=row_count, column=5).font = font
    ws.cell(row=row_count, column=6).font = font
    ws.cell(row=row_count, column=7).font = font
    ws.cell(row=row_count, column=8).font = font



def get_item_description_value(ws, row_count, result):
    ws.cell(row=row_count, column=3).value = result['item_name']
    ws.cell(row=row_count, column=4).value = result['mrp']
    ws.cell(row=row_count, column=5).value = result['quantity']
    ws.cell(row=row_count, column=6).value = result['mrp_total']
    ws.cell(row=row_count, column=7).value = result['discount']
    ws.cell(row=row_count, column=8).value = result['total_amount']

    ws.cell(row=row_count, column=3).font = font
    ws.cell(row=row_count, column=4).font = font
    ws.cell(row=row_count, column=5).font = font
    ws.cell(row=row_count, column=6).font = font
    ws.cell(row=row_count, column=7).font = font
    ws.cell(row=row_count, column=8).font = font


fill_type = ['darkHorizontal', 'gray0625', 'darkTrellis', 'lightUp', 'mediumGray', 'darkVertical', 'darkDown',
             'lightDown', 'lightTrellis', 'darkGray', 'darkGrid', 'darkUp', 'gray125', 'lightGrid', 'lightHorizontal',
             'lightVertical', 'solid', 'lightGray']


def get_item_return_value(ws, row_count, result):
    ws.cell(row=row_count, column=3).value = result['item_name']
    ws.cell(row=row_count, column=3).alignment = Alignment(wrap_text=True)
    ws.cell(row=row_count, column=4).value = result['mrp']
    ws.cell(row=row_count, column=5).value = result['quantity']
    ws.cell(row=row_count, column=6).value = result['mrp_total']
    ws.cell(row=row_count, column=7).value = result['discount']
    ws.cell(row=row_count, column=8).value = result['total_amount']

    ws.cell(row=row_count, column=4).font = font
    ws.cell(row=row_count, column=5).font = font
    ws.cell(row=row_count, column=6).font = font
    ws.cell(row=row_count, column=7).font = font
    ws.cell(row=row_count, column=8).font = font


    ws.cell(row=row_count, column=4).fill = PatternFill(fgColor="00FFFF00", fill_type="solid")
    ws.cell(row=row_count, column=5).fill = PatternFill(fgColor="00FFFF00", fill_type="solid")
    ws.cell(row=row_count, column=6).fill = PatternFill(fgColor="00FFFF00", fill_type="solid")
    ws.cell(row=row_count, column=7).fill = PatternFill(fgColor="00FFFF00", fill_type="solid")
    ws.cell(row=row_count, column=8).fill = PatternFill(fgColor="00FFFF00", fill_type="solid")

    thin = Side(border_style="thin", color="dddddd")
    ws.cell(row=row_count, column=4).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.cell(row=row_count, column=5).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.cell(row=row_count, column=6).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.cell(row=row_count, column=7).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.cell(row=row_count, column=8).border = Border(top=thin, left=thin, right=thin, bottom=thin)


def merge_row(ws, start_row, end_row):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row,
                   end_column=1)
    ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row,
                   end_column=2)
    ws.merge_cells(start_row=start_row, start_column=9, end_row=end_row,
                   end_column=9)
    ws.merge_cells(start_row=start_row, start_column=10, end_row=end_row,
                   end_column=10)
    ws.merge_cells(start_row=start_row, start_column=11, end_row=end_row,
                   end_column=11)
    ws.merge_cells(start_row=start_row, start_column=12, end_row=end_row,
                   end_column=12)
    ws.merge_cells(start_row=start_row, start_column=13, end_row=end_row,
                   end_column=13)
    ws.merge_cells(start_row=start_row, start_column=14, end_row=end_row,
                   end_column=14)
    ws.merge_cells(start_row=start_row, start_column=15, end_row=end_row,
                   end_column=15)
    ws.merge_cells(start_row=start_row, start_column=16, end_row=end_row,
                   end_column=16)


def get_others_value(ws, row_count, result, invoice_total):
    total = result['rounded_total'] + (0 if invoice_total['total_return_amount'] == '' else invoice_total['total_return_amount'])
    ws.cell(row=row_count, column=9).value = result['served_by']
    ws.cell(row=row_count, column=10).value = result['set_warehouse']
    ws.cell(row=row_count, column=11).value = result['special_discount']
    ws.cell(row=row_count, column=12).value = invoice_total['rounding_adjustment']
    ws.cell(row=row_count, column=13).value = result['total_taxes_and_charges']
    ws.cell(row=row_count, column=14).value = result['rounded_total']
    ws.cell(row=row_count, column=15).value = invoice_total['total_return_amount']
    ws.cell(row=row_count, column=16).value = total

    ws.cell(row=row_count, column=9).alignment = Alignment(vertical='top')
    ws.cell(row=row_count, column=10).alignment = Alignment(vertical='top')
    ws.cell(row=row_count, column=11).alignment = Alignment(vertical='top', horizontal='right')
    ws.cell(row=row_count, column=12).alignment = Alignment(vertical='top', horizontal='right')
    ws.cell(row=row_count, column=13).alignment = Alignment(vertical='top', horizontal='right')
    ws.cell(row=row_count, column=14).alignment = Alignment(vertical='top', horizontal='right')
    ws.cell(row=row_count, column=15).alignment = Alignment(vertical='top', horizontal='right')
    ws.cell(row=row_count, column=16).alignment = Alignment(vertical='top', horizontal='right')

    ws.cell(row=row_count, column=9).font = font
    ws.cell(row=row_count, column=10).font = font
    ws.cell(row=row_count, column=11).font = font
    ws.cell(row=row_count, column=12).font = font
    ws.cell(row=row_count, column=13).font = font
    ws.cell(row=row_count, column=14).font = font
    ws.cell(row=row_count, column=15).font = font
    ws.cell(row=row_count, column=16).font = font
    if (result['special_discount']):
        ws.cell(row=row_count, column=11).fill = PatternFill(fgColor="00FFFF00", fill_type="solid")


def get_total_row_value(ws, row_count, invoice_item_total):
    ws.cell(row=row_count, column=3).value = "Total"
    ws.cell(row=row_count, column=5).value = invoice_item_total['total_qty']
    ws.cell(row=row_count, column=6).value = invoice_item_total['mrp_total_price']
    ws.cell(row=row_count, column=7).value = invoice_item_total['total_discount']
    ws.cell(row=row_count, column=8).value = invoice_item_total['total_amount_']

    ws.cell(row=row_count, column=3).font = font
    ws.cell(row=row_count, column=5).font = font
    ws.cell(row=row_count, column=6).font = font
    ws.cell(row=row_count, column=7).font = font
    ws.cell(row=row_count, column=8).font = font
    ws.merge_cells(start_row=row_count, start_column=3, end_row=row_count,
                   end_column=4)


@frappe.whitelist()
def export_excel(filters):
    generate_excel_data(filters)
    site_name = frappe.local.site
    cur_dir = os.getcwd()
    save_folder_path = os.path.join(cur_dir, site_name, 'public/files/')
    pos_sales_data = []
    wb = Workbook()
    ws = wb.active
    data, query_result = get_invoice_data(filters)
    extra_row = 1
    thin = Side(border_style="thin", color="dddddd")
    ws.cell(row=extra_row, column=1).value = 'SL'
    ws.cell(row=extra_row, column=2).value = 'Invoice No/Customer'
    ws.cell(row=extra_row, column=3).value = 'Description'
    ws.cell(row=extra_row, column=9).value = 'Sales Person'
    ws.cell(row=extra_row, column=10).value = 'Store Name'
    ws.cell(row=extra_row, column=11).value = 'Special Discount'
    ws.cell(row=extra_row, column=12).value = 'Round'
    ws.cell(row=extra_row, column=13).value = 'VAT'
    ws.cell(row=extra_row, column=14).value = 'Total Amount'
    ws.cell(row=extra_row, column=15).value = 'Exchange'
    ws.cell(row=extra_row, column=16).value = 'Total'
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=8)

    for index, result in enumerate(query_result):
        sl_number = len(pos_sales_data) + 1
        next_invoice = True
        invoice_name = result.get('name')
        try:
            next_invoice = query_result[index + 1]
            next_invoice = query_result[index + 1]['name']
        except:
            next_invoice = not next_invoice

        invoice_item_total = data.get(invoice_name)
        if invoice_item_total is None:
            invoice_item_total = data.get(result['return_against'])

        time_obj = datetime.strptime(f"{result['posting_time']}", '%H:%M:%S.%f').strftime("%I:%M %p")
        style = 'style="background:#ffff0085"' if result['special_discount'] > 0 else ''

        row_count = index + extra_row + 1
        # print(invoice_item_total['exchange_item'], result['idx'], result['return_against'], invoice_name,
        #       result['is_return'])
        # if invoice_name == 'BR-20220911119995':
        #     break
        invoice_with_customer_info = invoice_name + '\n' + result['customer_name'] + '(' + result[
            'customer_mobile_number'] + ')' + '\n' + str(result['posting_date']) + ' ' + str(time_obj)

        if next_invoice and invoice_name == next_invoice:
            # print('invoice_name ', invoice_name)
            if invoice_name in pos_sales_data:
                if result['is_return'] == 1:
                    get_item_return_value(ws, row_count, result)
                else:
                    get_item_description_value(ws, row_count, result)
            elif result['return_against'] in pos_sales_data:
                if invoice_item_total['exchange_item'] == result['idx']:
                    ws.cell(row=row_count, column=3).value = 'Exchange Product List'
                    ws.cell(row=row_count, column=3).font = font
                    ws.cell(row=row_count, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                    ws.cell(row=row_count, column=3).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.merge_cells(start_row=row_count, start_column=3, end_row=row_count, end_column=8)
                    extra_row += 1
                    get_item_return_value(ws, row_count + 1, result)
                else:
                    get_item_return_value(ws, row_count, result)
            else:
                if result['is_return'] == 1:
                    ws.cell(row=row_count, column=1).value = sl_number
                    ws.cell(row=row_count, column=2).value = invoice_with_customer_info
                    ws.cell(row=row_count, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=row_count, column=2).alignment = Alignment(horizontal='center', vertical='center')
                    get_others_value(ws, row_count, result, invoice_item_total)
                    get_item_description_header(ws, row_count)
                    merge_row(ws, row_count, invoice_item_total['total_item_row'] + row_count - 1)
                    extra_row += 1
                    ws.cell(row=row_count + 1, column=3).value = 'Exchange Product List'
                    ws.cell(row=row_count + 1, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                    ws.cell(row=row_count + 1, column=3).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.cell(row=row_count + 1, column=3).alignment = Alignment(horizontal='center')
                    ws.merge_cells(start_row=row_count + 1, start_column=3, end_row=row_count + 1, end_column=8)
                    extra_row += 1
                    get_item_return_value(ws, row_count + 2, result)
                    pos_sales_data.append(invoice_name)

                    # elif result['return_against'] in pos_sales_data:
                    #     if invoice_item_total['exchange_item'] == result['idx']:
                    #         ws.cell(row=row_count, column=3).value = 'Exchange Product List'
                    #         ws.merge_cells(start_row=row_count, start_column=3, end_row=row_count, end_column=8)
                    #         get_item_return_value(ws, row_count + 1, result)
                    #     else:
                    #         get_item_return_value(ws, row_count, result)
                    # else:
                    # pos_sales_data[invoice_name] = [tr_return_with_header]
                    # get_item_description_header(ws, row_count)
                    # ws.merge_cells(start_row=row_count, start_column=1,
                    #                end_row=invoice_item_total['total_item_row'],
                    #                end_column=2)
                    # ws.merge_cells(start_row=row_count, start_column=9,
                    #                end_row=invoice_item_total['total_item_row'],
                    #                end_column=12)
                    # extra_row += 1
                    # get_item_return_value(ws, row_count + 1, result)
                else:
                    ws.cell(row=row_count, column=1).value = sl_number
                    ws.cell(row=row_count, column=2).value = invoice_with_customer_info
                    ws.cell(row=row_count, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=row_count, column=2).alignment = Alignment(horizontal='center', vertical='center')
                    merge_row(ws, row_count, invoice_item_total['total_item_row'] + row_count - 1)
                    get_item_description_header(ws, row_count)
                    extra_row += 1
                    get_item_description_value(ws, row_count + 1, result)
                    get_others_value(ws, row_count, result, invoice_item_total)
                    pos_sales_data.append(invoice_name)



        elif next_invoice and invoice_name != next_invoice:
            if invoice_name in pos_sales_data:
                if result['is_return'] == 1:
                    get_item_return_value(ws, row_count, result)
                    if result['special_note']:
                        extra_row += 1
                        ws.cell(row=row_count + 1, column=3).value = "Note:" + result['special_note']
                        ws.cell(row=row_count + 1, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                        ws.cell(row=row_count + 1, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                             bottom=thin)
                        ws.merge_cells(start_row=row_count + 1, start_column=3, end_row=row_count + 1,
                                       end_column=8)
                if result['is_return'] == 0:
                    get_item_description_value(ws, row_count, result)
                    extra_row += 1
                    get_total_row_value(ws, row_count + 1, invoice_item_total)
                    if result['special_note']:
                        extra_row += 1
                        ws.cell(row=row_count + 2, column=3).value = "Note:" + result['special_note']
                        ws.cell(row=row_count + 2, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                        ws.merge_cells(start_row=row_count + 2, start_column=3, end_row=row_count + 2,
                                       end_column=8)
                        ws.cell(row=row_count + 2, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                             bottom=thin)

            elif result['return_against'] in pos_sales_data:
                if invoice_item_total['exchange_item'] == result['idx']:
                    ws.cell(row=row_count, column=3).value = 'Exchange Product List'
                    ws.cell(row=row_count, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                    ws.cell(row=row_count + 3, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                         bottom=thin)
                    ws.cell(row=row_count, column=3).alignment = Alignment(horizontal='center')
                    ws.merge_cells(start_row=row_count, start_column=3, end_row=row_count, end_column=8)
                    extra_row += 1
                    get_item_return_value(ws, row_count + 1, result)
                    if result['special_note']:
                        extra_row += 1
                        ws.cell(row=row_count + 2, column=3).value = "Note:" + result['special_note']
                        ws.cell(row=row_count + 2, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                        ws.cell(row=row_count + 2, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                             bottom=thin)
                        ws.merge_cells(start_row=row_count + 2, start_column=3, end_row=row_count + 2,
                                       end_column=8)
                else:
                    get_item_return_value(ws, row_count, result)
                    if result['special_note']:
                        extra_row += 1
                        ws.cell(row=row_count + 1, column=3).value = "Note:" + result['special_note']
                        ws.cell(row=row_count + 1, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                        ws.cell(row=row_count + 1, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                             bottom=thin)
                        ws.merge_cells(start_row=row_count + 1, start_column=3, end_row=row_count + 1,
                                       end_column=8)
            else:
                if result['is_return'] == 1:
                    # if pos_sales_data == []:
                    #     if invoice_item_total['exchange_item'] == result['idx']:
                    #         ws.cell(row=row_count, column=1).value = sl_number
                    #         ws.cell(row=row_count, column=2).value = invoice_with_customer_info
                    #         ws.cell(row=row_count, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    #         ws.cell(row=row_count, column=2).alignment = Alignment(horizontal='center', vertical='center')
                    #         get_item_description_header(ws, row_count)
                    #         ws.merge_cells(start_row=row_count, start_column=1,
                    #                        end_row=invoice_item_total['total_item_row'] + row_count - 1,
                    #                        end_column=2)
                    #         ws.merge_cells(start_row=row_count, start_column=9,
                    #                        end_row=invoice_item_total['total_item_row'] + row_count - 1,
                    #                        end_column=12)
                    #         extra_row += 1
                    #         ws.cell(row=row_count + 1, column=3).value = 'Exchange Product List'
                    #         ws.cell(row=row_count + 1, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                    #         ws.cell(row=row_count + 1, column=3).border = Border(top=thin, left=thin, right=thin,
                    #                                                          bottom=thin)
                    #         ws.cell(row=row_count + 1, column=3).alignment = Alignment(horizontal='center', vertical='center')
                    #         ws.merge_cells(start_row=row_count + 1, start_column=3, end_row=row_count + 1, end_column=8)
                    #         extra_row += 1
                    #         get_item_return_value(ws, row_count + 2, result)
                    #         merge_row(ws, row_count + 2, invoice_item_total['total_item_row'])
                    #         if result['special_note']:
                    #             extra_row += 1
                    #             ws.cell(row=row_count + 3, column=3).value = "Note:" + result['special_note']
                    #             ws.cell(row=row_count + 3, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                    #             ws.cell(row=row_count + 3, column=3).border = Border(top=thin, left=thin, right=thin,
                    #                                                              bottom=thin)
                    #             ws.merge_cells(start_row=row_count + 3, start_column=3, end_row=row_count + 3,
                    #                            end_column=8)
                    #         pos_sales_data.append(invoice_name)

                    # if result['return_against'] in pos_sales_data:
                    #     if invoice_item_total['exchange_item'] == result['idx']:
                    #         ws.cell(row=row_count, column=3).value = 'Exchange Product List'
                    #         ws.cell(row=row_count, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                    #         ws.cell(row=row_count + 3, column=3).border = Border(top=thin, left=thin, right=thin,
                    #                                                              bottom=thin)
                    #         ws.cell(row=row_count, column=3).alignment = Alignment(horizontal='center', vertical='center')
                    #         ws.merge_cells(start_row=row_count, start_column=3, end_row=row_count, end_column=8)
                    #         extra_row += 1
                    #         get_item_return_value(ws, row_count + 1, result)
                    #         if result['special_note']:
                    #             extra_row += 1
                    #             ws.cell(row=row_count + 2, column=3).value = "Note:" + result['special_note']
                    #             ws.cell(row=row_count + 2, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                    #             ws.cell(row=row_count + 2, column=3).border = Border(top=thin, left=thin, right=thin,
                    #                                                              bottom=thin)
                    #             ws.merge_cells(start_row=row_count + 2, start_column=3, end_row=row_count + 2,
                    #                            end_column=8)
                    #     else:
                    #         get_item_return_value(ws, row_count, result)
                    #         if result['special_note']:
                    #             extra_row += 1
                    #             ws.cell(row=row_count + 1, column=3).value = "Note:" + result['special_note']
                    #             ws.cell(row=row_count + 1, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                    #             ws.cell(row=row_count + 1, column=3).border = Border(top=thin, left=thin, right=thin,
                    #                                                              bottom=thin)
                    #             ws.merge_cells(start_row=row_count + 1, start_column=3, end_row=row_count + 2,
                    #                            end_column=8)
                    # else:
                    ws.cell(row=row_count, column=1).value = sl_number
                    ws.cell(row=row_count, column=2).value = invoice_with_customer_info
                    ws.cell(row=row_count, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=row_count, column=2).alignment = Alignment(horizontal='center', vertical='center')

                    get_item_description_header(ws, row_count)
                    ws.merge_cells(start_row=row_count, start_column=1,
                                   end_row=invoice_item_total['total_item_row'],
                                   end_column=2)
                    ws.merge_cells(start_row=row_count, start_column=9,
                                   end_row=invoice_item_total['total_item_row'],
                                   end_column=12)
                    extra_row += 1
                    ws.cell(row=row_count + 1, column=3).value = 'Exchange Product List'
                    ws.cell(row=row_count + 1, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                    ws.cell(row=row_count + 1, column=3).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.cell(row=row_count + 1, column=3).alignment = Alignment(horizontal='center')
                    ws.merge_cells(start_row=row_count + 1, start_column=3, end_row=row_count + 1, end_column=8)
                    extra_row += 1
                    get_item_return_value(ws, row_count + 2, result)
                    merge_row(ws, row_count, invoice_item_total['total_item_row'])
                else:
                    # print('invoice_name => ', invoice_name, invoice_item_total['total_item_row'], row_count, invoice_item_total['total_item_row'] + row_count - 1)
                    ws.cell(row=row_count, column=1).value = sl_number
                    ws.cell(row=row_count, column=2).value = invoice_with_customer_info
                    ws.cell(row=row_count, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=row_count, column=2).alignment = Alignment(horizontal='center', vertical='center')

                    get_item_description_header(ws, row_count)
                    # ws.merge_cells(start_row=row_count, start_column=1,
                    #                end_row=invoice_item_total['total_item_row'],
                    #                end_column=1)
                    # ws.merge_cells(start_row=row_count, start_column=9,
                    #                end_row=invoice_item_total['total_item_row'],
                    #                end_column=12)
                    extra_row += 1
                    get_item_description_value(ws, row_count + 1, result)
                    extra_row += 1
                    get_total_row_value(ws, row_count + 2, invoice_item_total)
                    if result['special_note']:
                        extra_row += 1
                        ws.cell(row=row_count + 3, column=3).value = "Note:" + result['special_note']
                        ws.cell(row=row_count + 3, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                        ws.cell(row=row_count + 3, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                             bottom=thin)
                        ws.merge_cells(start_row=row_count + 3, start_column=3, end_row=row_count + 3, end_column=8)
                    #     merge_row(ws, row_count, invoice_item_total['total_item_row'] + row_count - 1)
                    # else:
                    merge_row(ws, row_count, invoice_item_total['total_item_row'] + row_count - 1)
                    get_others_value(ws, row_count, result, invoice_item_total)
                    pos_sales_data.append(invoice_name)
        elif next_invoice == False:
            # if invoice_name in pos_sales_data:
            #     if result['is_return'] == 1:
            #         get_item_return_value(ws, row_count, result)
            #     # else:
            #     #     get_total_row_value(ws, row_count, result)
            #     if result['is_return'] == 0:
            #         get_item_description_value(ws, row_count, result)
            #         extra_row += 1
            #         get_total_row_value(ws, row_count + 1, invoice_item_total)
            #     if result['special_note']:
            #         extra_row += 1
            #         ws.cell(row=row_count + 2, column=3).value = "Note:" + result['special_note']
            #         ws.cell(row=row_count + 2, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
            #         ws.cell(row=row_count + 2, column=3).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            #         ws.merge_cells(start_row=row_count + 2, start_column=3, end_row=row_count + 2, end_column=8)

            if invoice_name in pos_sales_data:
                if result['is_return'] == 1:
                    get_item_return_value(ws, row_count, result)
                    if result['special_note']:
                        extra_row += 1
                        ws.cell(row=row_count + 1, column=3).value = "Note:" + result['special_note']
                        ws.cell(row=row_count + 1, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                        ws.cell(row=row_count + 1, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                             bottom=thin)
                        ws.merge_cells(start_row=row_count + 1, start_column=3, end_row=row_count + 1,
                                       end_column=8)
                if result['is_return'] == 0:
                    get_item_description_value(ws, row_count, result)
                    extra_row += 1
                    get_total_row_value(ws, row_count + 1, invoice_item_total)
                    if result['special_note']:
                        extra_row += 1
                        ws.cell(row=row_count + 2, column=3).value = "Note:" + result['special_note']
                        ws.cell(row=row_count + 2, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                        ws.merge_cells(start_row=row_count + 2, start_column=3, end_row=row_count + 2,
                                       end_column=8)
                        ws.cell(row=row_count + 2, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                             bottom=thin)

            elif result['return_against'] in pos_sales_data:
                if invoice_item_total['exchange_item'] == result['idx']:
                    ws.cell(row=row_count, column=3).value = 'Exchange Product List'
                    ws.cell(row=row_count, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                    ws.cell(row=row_count, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                     bottom=thin)
                    ws.cell(row=row_count, column=3).alignment = Alignment(horizontal='center')
                    ws.merge_cells(start_row=row_count, start_column=3, end_row=row_count, end_column=8)
                    extra_row += 1
                    get_item_return_value(ws, row_count + 1, result)
                    if result['special_note']:
                        extra_row += 1
                        ws.cell(row=row_count + 2, column=3).value = "Note:" + result['special_note']
                        ws.cell(row=row_count + 2, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                        ws.merge_cells(start_row=row_count + 2, start_column=3, end_row=row_count + 2,
                                       end_column=8)
                        ws.cell(row=row_count + 2, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                             bottom=thin)
                else:
                    get_item_return_value(ws, row_count, result)
                    if result['special_note']:
                        extra_row += 1
                        ws.cell(row=row_count + 1, column=3).value = "Note:" + result['special_note']
                        ws.cell(row=row_count + 1, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                        ws.cell(row=row_count + 1, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                             bottom=thin)
                        ws.merge_cells(start_row=row_count + 1, start_column=3, end_row=row_count + 1,
                                       end_column=8)

            else:
                if result['is_return'] == 1:
                    if not pos_sales_data:
                        if invoice_item_total['exchange_item'] == result['idx']:
                            ws.cell(row=row_count, column=1).value = sl_number
                            ws.cell(row=row_count, column=2).value = invoice_with_customer_info

                            ws.cell(row=row_count, column=1).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')
                            ws.cell(row=row_count, column=2).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')
                            get_item_description_header(ws, row_count)
                            ws.merge_cells(start_row=row_count, start_column=1,
                                           end_row=invoice_item_total['total_item_row'],
                                           end_column=2)
                            ws.merge_cells(start_row=row_count, start_column=9,
                                           end_row=invoice_item_total['total_item_row'],
                                           end_column=12)
                            extra_row += 1
                            ws.cell(row=row_count + 1, column=3).value = 'Exchange Product List'
                            ws.cell(row=row_count + 1, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                            ws.cell(row=row_count + 1, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                                 bottom=thin)
                            ws.cell(row=row_count + 1, column=3).alignment = Alignment(horizontal='center')
                            ws.merge_cells(start_row=row_count + 1, start_column=3, end_row=row_count + 1, end_column=8)
                            extra_row += 1
                            get_item_return_value(ws, row_count + 2, result)
                            merge_row(ws, row_count + 2, invoice_item_total['total_item_row'])

                            if result['special_note']:
                                extra_row += 1
                                ws.cell(row=row_count + 3, column=3).value = "Note:" + result['special_note']
                                ws.cell(row=row_count + 3, column=3).fill = PatternFill(fgColor="a9d18e",
                                                                                        fill_type="solid")
                                ws.cell(row=row_count + 3, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                                     bottom=thin)
                                ws.merge_cells(start_row=row_count + 3, start_column=3, end_row=row_count + 3,
                                               end_column=8)
                            pos_sales_data.append(invoice_name)

                        else:
                            get_item_return_value(ws, row_count, result)
                            if result['special_note']:
                                extra_row += 1
                                ws.cell(row=row_count + 1, column=3).value = "Note:" + result['special_note']
                                ws.cell(row=row_count + 1, column=3).fill = PatternFill(fgColor="a9d18e",
                                                                                        fill_type="solid")
                                ws.merge_cells(start_row=row_count + 1, start_column=3, end_row=row_count + 1,
                                               end_column=8)
                                ws.cell(row=row_count + 1, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                                     bottom=thin)

                    else:
                        # pos_sales_data[invoice_name] = [tr_return_with_header]
                        ws.cell(row=row_count, column=1).value = sl_number
                        ws.cell(row=row_count, column=2).value = invoice_with_customer_info

                        ws.cell(row=row_count, column=1).alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(row=row_count, column=2).alignment = Alignment(horizontal='center', vertical='center')
                        get_item_description_header(ws, row_count)
                        ws.merge_cells(start_row=row_count, start_column=1,
                                       end_row=invoice_item_total['total_item_row'],
                                       end_column=2)
                        ws.merge_cells(start_row=row_count, start_column=9,
                                       end_row=invoice_item_total['total_item_row'],
                                       end_column=12)
                        ws.cell(row=row_count, column=3).value = 'Exchange Product List'
                        ws.cell(row=row_count, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                        ws.cell(row=row_count, column=3).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        ws.cell(row=row_count, column=3).alignment = Alignment(horizontal='center')
                        ws.merge_cells(start_row=row_count, start_column=3, end_row=row_count, end_column=8)
                        extra_row += 1
                        get_item_return_value(ws, row_count + 1, result)
                        merge_row(ws, row_count, invoice_item_total['total_item_row'])

                        if result['special_note']:
                            extra_row += 1
                            ws.cell(row=row_count + 2, column=3).value = "Note:" + result['special_note']
                            ws.cell(row=row_count + 2, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                            ws.merge_cells(start_row=row_count + 2, start_column=3, end_row=row_count + 2,
                                           end_column=8)
                            ws.cell(row=row_count + 2, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                                 bottom=thin)
                        pos_sales_data.append(invoice_name)
                else:
                    # pos_sales_data[invoice_name] = [tr_first_row]
                    # pos_sales_data[invoice_name].append(tr_total_row)
                    ws.cell(row=row_count, column=1).value = sl_number
                    ws.cell(row=row_count, column=2).value = invoice_with_customer_info
                    ws.cell(row=row_count, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=row_count, column=2).alignment = Alignment(horizontal='center', vertical='center')
                    get_item_description_header(ws, row_count)
                    # ws.merge_cells(start_row=row_count, start_column=1,
                    #                end_row=invoice_item_total['total_item_row'],
                    #                end_column=1)
                    # ws.merge_cells(start_row=row_count, start_column=9,
                    #                end_row=invoice_item_total['total_item_row'],
                    #                end_column=12)
                    extra_row += 1
                    get_item_description_value(ws, row_count + 1, result)
                    extra_row += 1
                    get_total_row_value(ws, row_count + 2, invoice_item_total)
                    if result['special_note']:
                        extra_row += 1
                        ws.cell(row=row_count + 3, column=3).value = "Note:" + result['special_note']
                        ws.cell(row=row_count + 3, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                        ws.cell(row=row_count + 3, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                             bottom=thin)
                        ws.merge_cells(start_row=row_count + 3, start_column=3, end_row=row_count + 3, end_column=8)
                        merge_row(ws, row_count, invoice_item_total['total_item_row'] + row_count)
                    else:
                        merge_row(ws, row_count, invoice_item_total['total_item_row'] + row_count - 1)

                    get_others_value(ws, row_count, result, invoice_item_total)
                    pos_sales_data.append(invoice_name)

        # print(pos_sales_data)
        #
        # tr_first_row = f"""
        #     <tr>
        #         <th scope='row' rowspan='{invoice_item_total['total_item_row']}'>{sl_number}</th>
        #         <td rowspan='{invoice_item_total['total_item_row']}' class="text-center">{invoice_name}<br>{result['customer_name']} ({result['customer_mobile_number']})<br>{result['posting_date']} {time_obj}</td>
        #         <td>Name</td>
        #         <td>MRP</td>
        #         <td>Qty</td>
        #         <td>MRP Total</td>
        #         <td>Disc</td>
        #         <td>Total</td>
        #         <td rowspan='{invoice_item_total['total_item_row']}'>{result['served_by']}</td>
        #         <td rowspan='{invoice_item_total['total_item_row']}'>{result['set_warehouse']}</td>
        #         <td {style} rowspan='{invoice_item_total['total_item_row']}'>{result['special_discount']}</td>
        #         <td rowspan='{invoice_item_total['total_item_row']}'>{result['rounding_adjustment']}</td>
        #         <td rowspan='{invoice_item_total['total_item_row']}'>{result['total_taxes_and_charges']}</td>
        #         <td rowspan='{invoice_item_total['total_item_row']}'>{result['net_total']}</td>
        #         <td rowspan='{invoice_item_total['total_item_row']}'>{invoice_item_total['total_return_amount']}</td>
        #         <td rowspan='{invoice_item_total['total_item_row']}'>{result['rounded_total']}</td>
        #     </tr>
        #
        #     <tr>
        #         <td>{result.get('item_name')}</td>
        #         <td>{result.get('mrp')}</td>
        #         <td>{result.get('quantity')}</td>
        #         <td>{result.get('mrp_total')}</td>
        #         <td>{result.get('discount')}</td>
        #         <td>{result.get('total_amount')}</td>
        #     </tr>
        # """
        #
        # tr_return_with_header = f"""
        #     <tr>
        #         <th scope='row' rowspan='{invoice_item_total['total_item_row']}'>{sl_number}</th>
        #         <td rowspan='{invoice_item_total['total_item_row']}' class="text-center">{invoice_name}<br>{result['customer_name']} ({result['customer_mobile_number']})<br>{result['posting_date']} {time_obj}</td>
        #         <td>Name</td>
        #         <td>MRP</td>
        #         <td>Qty</td>
        #         <td>MRP Total</td>
        #         <td>Disc</td>
        #         <td>Total</td>
        #         <td rowspan='{invoice_item_total['total_item_row']}'>{result['served_by']}</td>
        #         <td rowspan='{invoice_item_total['total_item_row']}'>{result['set_warehouse']}</td>
        #         <td {style} rowspan='{invoice_item_total['total_item_row']}'>{result['special_discount']}</td>
        #         <td rowspan='{invoice_item_total['total_item_row']}'>{invoice_item_total['rounding_adjustment']}</td>
        #         <td rowspan='{invoice_item_total['total_item_row']}'>{result['total_taxes_and_charges']}</td>
        #         <td rowspan='{invoice_item_total['total_item_row']}'>{result['total_amount']}</td>
        #         <td rowspan='{invoice_item_total['total_item_row']}'>{invoice_item_total['total_return_amount']}</td>
        #         <td rowspan='{invoice_item_total['total_item_row']}'>{result['rounded_total']}</td>
        #     </tr>
        #     <tr>
        #         <td colspan='6' class="font-weight-bold text-center" style="background:#c3e0aad4">Exchange Product List</td>
        #     </tr>
        #     <tr>
        #         <td>{result.get('item_name')}</td>
        #         <td style="background:#ffff0085">{result.get('mrp')}</td>
        #         <td style="background:#ffff0085">{result.get('quantity')}</td>
        #         <td style="background:#ffff0085">{result.get('selling_rate')}</td>
        #         <td style="background:#ffff0085">{result.get('discount')}</td>
        #         <td style="background:#ffff0085">{result.get('total_amount')}</td>
        #     </tr>
        # """
        #
        # tr_rowspan = f"""
        #     <tr>
        #         <td>{result.get('item_name')}</td>
        #         <td>{result.get('mrp')}</td>
        #         <td>{result.get('quantity')}</td>
        #         <td>{result.get('mrp_total')}</td>
        #         <td>{result.get('discount')}</td>
        #         <td>{result.get('total_amount')}</td>
        #     </tr>
        # """
        #
        # tr_total_row = f"""
        #     <tr>
        #         <td colspan='2'>Total</td>
        #         <td>{invoice_item_total.get('total_qty')}</td>
        #         <td>{invoice_item_total.get('mrp_total_price')}</td>
        #         <td>{invoice_item_total.get('total_discount')}</td>
        #         <td>{invoice_item_total.get('total_amount_')}</td>
        #     </tr>
        # """
        #
        # tr_special_note = f"""
        #     <tr style="background:#c3e0aad4">
        #         <td colspan='6'>Note: {result['special_note']}</td>
        #     </tr>
        # """
        # tr_return_head = """
        #     <tr>
        #         <td colspan='6' class="font-weight-bold text-center" style="background:#c3e0aad4">Exchange Product List</td>
        #     </tr>
        # """
        # tr_return_row = f"""
        #     <tr>
        #         <td>{result.get('item_name')}</td>
        #         <td style="background:#ffff0085">{result.get('mrp')}</td>
        #         <td style="background:#ffff0085">{result.get('quantity')}</td>
        #         <td style="background:#ffff0085">{result.get('mrp_total')}</td>
        #         <td style="background:#ffff0085">{result.get('discount')}</td>
        #         <td style="background:#ffff0085">{result.get('total_amount')}</td>
        #     </tr>
        # """
        #
        # if next_invoice and invoice_name == next_invoice:
        #     if sales_data.get(invoice_name):
        #         if result['is_return'] == 1:
        #             sales_data[invoice_name].append(tr_return_row)
        #         else:
        #             sales_data[invoice_name].append(tr_rowspan)
        #     else:
        #         if result['is_return'] == 1:
        #             if sales_data == {}:
        #                 sales_data[invoice_name] = [tr_return_with_header]
        #             elif sales_data.get(result['return_against']):
        #                 if invoice_item_total['exchange_item'] == result['idx']:
        #                     sales_data[result['return_against']].append(tr_return_head)
        #                 sales_data[result['return_against']].append(tr_return_row)
        #             else:
        #                 sales_data[invoice_name] = [tr_return_with_header]
        #         else:
        #             sales_data[invoice_name] = [tr_first_row]
        #
        # elif next_invoice and invoice_name != next_invoice:
        #     if sales_data.get(invoice_name):
        #         if result['is_return'] == 1:
        #             sales_data[invoice_name].append(tr_return_row)
        #         else:
        #             sales_data[invoice_name].append(tr_rowspan)
        #         if result['is_return'] == 0:
        #             sales_data[invoice_name].append(tr_total_row)
        #         if result['special_note']:
        #             sales_data[invoice_name].append(tr_special_note)
        #     else:
        #         if result['is_return'] == 1:
        #             if sales_data == {}:
        #                 if invoice_item_total['exchange_item'] == result['idx']:
        #                     sales_data[invoice_name].append(tr_return_head)
        #                 sales_data[invoice_name] = [tr_return_with_header]
        #                 if result['special_note']:
        #                     sales_data[invoice_name].append(tr_special_note)
        #             elif sales_data.get(result['return_against']):
        #                 if invoice_item_total['exchange_item'] == result['idx']:
        #                     sales_data[result['return_against']].append(tr_return_head)
        #                 sales_data[result['return_against']].append(tr_return_row)
        #                 if result['special_note']:
        #                     sales_data[result['return_against']].append(tr_special_note)
        #             else:
        #                 sales_data[invoice_name] = [tr_return_with_header]
        #             # if result['special_note']:
        #             #     sales_data[result['return_against']].append(tr_special_note)
        #         else:
        #             sales_data[invoice_name] = [tr_first_row]
        #             sales_data[invoice_name].append(tr_total_row)
        #             if result['special_note']:
        #                 sales_data[invoice_name].append(tr_special_note)
        #
        # elif next_invoice == False:
        #     if sales_data.get(invoice_name):
        #         if result['is_return'] == 1:
        #             sales_data[invoice_name].append(tr_return_row)
        #         else:
        #             sales_data[invoice_name].append(tr_rowspan)
        #         if result['is_return'] == 0:
        #             sales_data[invoice_name].append(tr_total_row)
        #         if result['special_note']:
        #             sales_data[invoice_name].append(tr_special_note)
        #
        #     else:
        #         if result['is_return'] == 1:
        #             if sales_data == {}:
        #                 if invoice_item_total['exchange_item'] == result['idx']:
        #                     sales_data[invoice_name].append(tr_return_head)
        #                 sales_data[invoice_name] = [tr_return_with_header]
        #                 if result['special_note']:
        #                     sales_data[invoice_name].append(tr_special_note)
        #             elif sales_data.get(result['return_against']):
        #                 if invoice_item_total['exchange_item'] == result['idx']:
        #                     sales_data[result['return_against']].append(tr_return_head)
        #                 sales_data[result['return_against']].append(tr_return_row)
        #                 if result['special_note']:
        #                     sales_data[result['return_against']].append(tr_special_note)
        #             else:
        #                 sales_data[invoice_name] = [tr_return_with_header]
        #             # sales_data[result['return_against']].append(tr_special_note)
        #         else:
        #             sales_data[invoice_name] = [tr_first_row]
        #             sales_data[invoice_name].append(tr_total_row)
        #             if result['special_note']:
        #                 sales_data[result['return_against']].append(tr_special_note)

    wb.save(save_folder_path + 'detail_sales_report.xlsx')
    url = urljoin(frappe.utils.get_url(), 'files/detail_sales_report.xlsx')
    return url


@frappe.whitelist()
def generate_excel_data(**filters):
    # filters = json.loads(kwargs)
    data, query_result = get_invoice_data(filters)
    sales_data = {}

    site_name = frappe.local.site
    cur_dir = os.getcwd()
    save_folder_path = os.path.join(cur_dir, site_name, 'public/files/')
    wb = Workbook()
    ws = wb.active
    extra_row = 1
    thin = Side(border_style="thin", color="dddddd")
    ws.cell(row=extra_row, column=1).value = 'SL'
    ws.cell(row=extra_row, column=2).value = 'Invoice No/Customer'
    ws.cell(row=extra_row, column=3).value = 'Description'
    ws.cell(row=extra_row, column=9).value = 'Sales Person'
    ws.cell(row=extra_row, column=10).value = 'Store Name'
    ws.cell(row=extra_row, column=11).value = 'Special Discount'
    ws.cell(row=extra_row, column=12).value = 'Round'
    ws.cell(row=extra_row, column=13).value = 'VAT'
    ws.cell(row=extra_row, column=14).value = 'Total Amount'
    ws.cell(row=extra_row, column=15).value = 'Exchange'
    ws.cell(row=extra_row, column=16).value = 'Total'
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=8)
    total_item_qty = 0
    total_exchange_qty = 0
    unit_total_amount = 0
    product_discount_amount = 0
    invoice_discount = 0
    sale_amount = 0
    total_exchange_amount = 0
    for index, result in enumerate(query_result):
        if result.get('is_return') == 1:
            total_exchange_qty += result.get('quantity')
        else:
            total_item_qty += result.get('quantity')

        unit_total_amount += result.get('mrp') * result.get('quantity')
        product_discount_amount += result.get('discount')

        next_invoice = True
        invoice_name = result.get('name')
        try:
            next_invoice = query_result[index + 1]
            next_invoice = query_result[index + 1]['name']
        except:
            next_invoice = not next_invoice

        invoice_item_total = data.get(invoice_name)
        if invoice_item_total is None:
            invoice_item_total = data.get(result['return_against'])

        if invoice_name in sales_data:
            if invoice_item_total['exchange_item'] == result['idx'] and result['is_return'] == 1:
                sales_data[invoice_name].append({'return_head': "Exchange Product List"})
                sales_data[invoice_name].append(result)
                if next_invoice and invoice_name != next_invoice or next_invoice == False:
                    sales_data[invoice_name].append(result)
                    if result['special_note']:
                        sales_data[invoice_name].append({'note': result['special_note']})
            else:
                sales_data[invoice_name].append(result)
                if next_invoice and invoice_name != next_invoice or next_invoice == False:
                    if result['is_return'] == 0:
                        sales_data[invoice_name].append({'total_row': True, 'name': result.get('name'),
                                                         'return_against': result.get('return_against', None)})
                    if result['special_note']:
                        sales_data[invoice_name].append({'note': result['special_note']})

        elif result['return_against'] in sales_data:
            if invoice_item_total['exchange_item'] == result['idx'] and result['is_return'] == 1:
                sales_data[result['return_against']].append({'return_head': "Exchange Product List"})
                sales_data[result['return_against']].append(result)
                if next_invoice and invoice_name != next_invoice or next_invoice == False:
                    if result['special_note']:
                        sales_data[result['return_against']].append({'note': result['special_note']})
            else:
                sales_data[result['return_against']].append(result)
                if next_invoice and invoice_name != next_invoice or next_invoice == False:
                    if result['special_note']:
                        sales_data[result['return_against']].append({'note': result['special_note']})
        else:
            if invoice_item_total['exchange_item'] == result['idx'] and result['is_return'] == 1:
                sales_data[invoice_name] = [{'return_head': "Exchange Product List"}]
                sales_data[invoice_name].append(result)
                if next_invoice and invoice_name != next_invoice or next_invoice == False:
                    sales_data[invoice_name].append(result)
                    if result['special_note']:
                        sales_data[invoice_name].append({'note': result['special_note']})
            else:
                if result['is_return'] == 1:
                    invoice_data = {}
                    invoice_data['return_head'] = "Exchange Product List"
                    invoice_data['posting_date'] = result['posting_date']
                    invoice_data['posting_time'] = result['posting_time']
                    invoice_data['customer_name'] = result['customer_name']
                    invoice_data['customer_mobile_number'] = result['customer_mobile_number']
                    invoice_data['served_by'] = result['served_by']
                    invoice_data['set_warehouse'] = result['set_warehouse']
                    invoice_data['special_discount'] = result['special_discount']
                    invoice_data['total_taxes_and_charges'] = result['total_taxes_and_charges']
                    invoice_data['total_amount'] = result['total_amount']
                    invoice_data['rounded_total'] = result['rounded_total']

                    sales_data[invoice_name] = [invoice_data]
                    sales_data[invoice_name].append(result)
                elif result['is_return'] == 0:
                    sales_data[invoice_name] = [result]
                if next_invoice and invoice_name != next_invoice or next_invoice == False:
                    if result['is_return'] == 0:
                        sales_data[invoice_name].append({'total_row': True, 'name': result.get('name'),
                                                         'return_against': result.get('return_against', None)})
                    if result['special_note']:
                        sales_data[invoice_name].append({'note': result['special_note']})

    row_count = 2
    for index, invoice_name in enumerate(sales_data):
        res = sales_data[invoice_name][0]
        invoice_item_total = data.get(invoice_name)
        if invoice_item_total is None:
            invoice_item_total = data.get(sales_data[invoice_name][0]['return_against'])

        time_obj = datetime.strptime(f"{res['posting_time']}", '%H:%M:%S.%f').strftime("%I:%M %p")
        invoice_with_customer_info = invoice_name + '\n' + res['customer_name'] + '(' + res[
            'customer_mobile_number'] + ')' + '\n' + str(res['posting_date']) + ' ' + str(time_obj)
        ws.cell(row=row_count, column=1).value = index + 1
        ws.cell(row=row_count, column=2).value = invoice_with_customer_info

        ws.cell(row=row_count, column=1).font = font
        ws.cell(row=row_count, column=2).font = font
        ws.cell(row=row_count, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_count, column=2).alignment = Alignment(horizontal='center', vertical='center')
        get_item_description_header(ws, row_count)
        get_others_value(ws, row_count, res, invoice_item_total)
        row_length = len(sales_data[invoice_name]) + row_count

        invoice_discount += res.get('special_discount', 0)
        total_exchange_amount += 0 if invoice_item_total.get('total_return_amount') == '' else invoice_item_total.get('total_return_amount')
        sale_amount += res.get('rounded_total', 0)

        merge_row(ws, row_count, row_length)
        row_count += 1

        for result in sales_data[invoice_name]:
            if result.get('return_head', None):
                ws.cell(row=row_count, column=3).value = 'Exchange Product List'
                ws.cell(row=row_count, column=3).font = font
                ws.cell(row=row_count, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                ws.cell(row=row_count, column=3).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(row=row_count, column=3).alignment = Alignment(horizontal='center')
                ws.merge_cells(start_row=row_count, start_column=3, end_row=row_count, end_column=8)
                row_count += 1
                continue
            if result.get('is_return', None) == 1:
                get_item_return_value(ws, row_count, result)
                row_count += 1
            elif result.get('total_row', None):
                get_total_row_value(ws, row_count, invoice_item_total)
                row_count += 1
            elif result.get('is_return', None) == 0:
                get_item_description_value(ws, row_count, result)
                row_count += 1
            if result.get('note', None):
                ws.cell(row=row_count, column=3).value = "Note : " + result['note']
                ws.cell(row=row_count, column=3).font = font
                ws.cell(row=row_count, column=3).fill = PatternFill(fgColor="a9d18e", fill_type="solid")
                ws.merge_cells(start_row=row_count, start_column=3, end_row=row_count,
                               end_column=8)
                ws.cell(row=row_count, column=3).border = Border(top=thin, left=thin, right=thin,
                                                                 bottom=thin)
                row_count += 1

    ws.cell(row=row_count, column=1).value = "Total"
    ws.cell(row=row_count, column=2).value = "Total Qty=" + str(total_item_qty) + '\n' + "Exchange Qty=" + str(
        total_exchange_qty)
    ws.cell(row=row_count, column=3).value = "Unit Total=" + str(unit_total_amount) + ', ' + "Product Disc=" + str(
        product_discount_amount)
    ws.cell(row=row_count, column=9).value = "Invoice Total=" + '\n' + str(unit_total_amount - product_discount_amount)
    ws.cell(row=row_count, column=10).value = "Special Disc=" + '\n' + str(invoice_discount)
    ws.cell(row=row_count, column=11).value = "Sale Amt=" + str(sale_amount)
    ws.cell(row=row_count, column=13).value = "Total Ex.=" + str(total_exchange_amount)
    ws.cell(row=row_count, column=15).value = "Org Sale.=" + str(sale_amount + total_exchange_amount)

    for column in range(1, 17):
        ws.merge_cells(start_row=row_count, start_column=column, end_row=row_count+1,
                       end_column=column)
    ws.merge_cells(start_row=row_count, start_column=3, end_row=row_count+1,
                   end_column=8)
    ws.merge_cells(start_row=row_count, start_column=11, end_row=row_count+1,
                   end_column=12)
    ws.merge_cells(start_row=row_count, start_column=13, end_row=row_count+1,
                   end_column=14)
    ws.merge_cells(start_row=row_count, start_column=15, end_row=row_count+1,
                   end_column=16)

    # column_widths = []
    # for row in data:
    #     for i, cell in enumerate(row):
    #         if len(column_widths) > i:
    #             if len(cell) > column_widths[i]:
    #                 column_widths[i] = len(cell)
    #         else:
    #             column_widths += [len(cell)]

    # for culumn_number in range(1, 17):  # ,1 to start at 1
    ws.column_dimensions[get_column_letter(2)].width = 18

    file = wb.save(save_folder_path + 'detail_sales_report_2.xlsx')
    file_path = save_folder_path + 'detail_sales_report_2.xlsx'
    # file_path = urljoin(frappe.utils.get_url(), 'files/detail_sales_report_2.xlsx')

    # byte_data = io.BytesIO()
    # wb.save(byte_data)
    # # The zip compressor
    # # zf = ZipFile(byte_data, "w")
    #
    # # for file in files:
    # #     zf.write(file, basename(file))
    #
    # # Must close zip for all contents to be written
    # # zf.close()



    with open(file_path, "rb") as file:
        frappe.local.response.filename = 'detail_sales_report_2.xlsx'
        frappe.local.response.filecontent = file.read()
        frappe.local.response.type = "download"

    return file_path

