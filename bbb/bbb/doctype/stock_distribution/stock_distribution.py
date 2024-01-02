# Copyright (c) 2023, invento software limited and contributors
# For license information, please see license.txt

import frappe
from frappe import _
import json
from frappe.model.document import Document
import os
import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import random
import string
import pandas as pd

class StockDistribution(Document):
    
    def validate(self):
        total_predict = 0
        if self.outlet_selection_table:
            for x in self.outlet_selection_table:
                if x.get("percentage"):
                    total_predict += x.get("percentage")
                self.send_system_notification_to_outlet_manager(x.get("warehouse"))
        if total_predict != 100:
            frappe.throw("Outlet Prediction total must be 100")
        
    def send_system_notification_to_outlet_manager(self,warehouse):
            outlet_manager = frappe.get_value("User",{"outlet" : warehouse},"name")
            if outlet_manager:
                notification = frappe.new_doc('Notification Log')
                notification.type = 'Alert'
                notification.document_type = self.doctype
                notification.document_name = self.name
                notification.subject = 'New Stock coming on {date}'.format(date= frappe.utils.formatdate(self.expected_delivery_date))
                notification.email_content = 'There Is a new stock comming from distribution <a href="/app/stock-distribution/{dis}" >{dis}</a>, to outlet {outlet}'.format(dis=self.name,outlet=warehouse)
                notification.from_user = frappe.session.user
                notification.for_user = outlet_manager
                notification.insert(ignore_permissions=True)
            
    def on_submit(self):
        if self.ignore_validation:
            self.stock_entry()
        else:
            if self.upload_distribution_excell and self.against_purchase_receipt:
                excel_file_path = get_site_directory_path() + self.upload_distribution_excell
                excell_dict = {}
                excell_item_list = []
                grand_total = 0
                message = "In Excell Qty Isn't Equal For Item "
                excell_data = []
                workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
                sheet = workbook.active
                labels = [cell.value for cell in sheet[1]]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    item_code = "{}".format(row[0])
                    excell_item_list.append(item_code)
                    total = 0
                    row_data = {}
                    for x in row[1:]:
                        total += x
                        grand_total += x
                    for i, value in enumerate(row):
                        label = labels[i]
                        row_data[label] = value
                    excell_data.append(row_data)
                    excell_dict[item_code] = total
                    
                purchase_receipt = frappe.get_doc("Purchase Receipt",self.against_purchase_receipt)
                if purchase_receipt.items:
                    for item in purchase_receipt.items:
                        if item.get("item_code") in excell_item_list:
                            excell_qty = excell_dict.get(item.get("item_code"))
                            pr_qty = item.get("qty")
                            if excell_qty != pr_qty:
                                remain = pr_qty - excell_qty
                                single_message = "<b>{item}</b> need <b>{qt_y}</b>\n, ".format(item=item.get("item_code"),qt_y = remain)
                                message += single_message
                        else:
                            frappe.throw("Item {} Isn't Available In Distribution Excell".format(item.get("item_code")))
                            
                if grand_total != purchase_receipt.total_qty:
                    frappe.throw(message)
                    
                self.create_stock_transfer(purchase_receipt.items,excell_data,purchase_receipt.name)
            else:
                frappe.throw("Please Upload Distribution excell And Set Against Purchase Receipt")
                
    def stock_entry(self):
        if self.upload_distribution_excell:
            excel_file_path = get_site_directory_path() + self.upload_distribution_excell
            excell_data = []
            workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
            sheet = workbook.active
            labels = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                item_code = "{}".format(row[0])
                total = 0
                row_data = {}
                for x in row[1:]:
                    if x:
                        total += x
                for i, value in enumerate(row):
                    label = labels[i]
                    row_data[label] = value
                excell_data.append(row_data)
        else:
            frappe.throw("Upload Distribution Excell")
            
        self.create_stock_transfer(self.purchase_distribution_items,excell_data)
        
    def create_stock_transfer(self,purchase_receipt,excell_data,purchase_receipt_reference = None):
        sd = self.name
        company = self.company
        items = []
        warehouses= []
        for item in purchase_receipt:
            for outlet in excell_data:
                if item.get("item_code") == str(outlet.get("Item Code")):
                    for key,value in outlet.items():
                        if key != "Item Code" and value and value > 0:
                            final = ""
                            warehouse = key.lower().replace("_"," ").replace("&","-").title()
                            up = warehouse.split("-")
                            final += up[0]
                            final += "-"
                            final += up[1].upper()
                            warehouses.append(final)
                             
                            # Create Stock Entry Data
                            data_dict = {}
                            data_dict["item_code"] = item.get("item_code")
                            data_dict["qty"] = value
                            data_dict["transfer_qty_from_stock_distribution"] = value
                            data_dict["s_warehouse"] = item.get("warehouse")
                            data_dict["t_warehouse"] = final
                            data_dict["uom"] = item.get("uom")
                            data_dict["reference_purchase_receipt"] = purchase_receipt_reference if purchase_receipt_reference else ""
                            items.append(data_dict)
        
        # frappe.msgprint((str(items)))
        for x in set(warehouses):
            update_items = []
            for y in items:
                if x == y.get("t_warehouse"):
                    update_items.append(y)
                    
            stock_entry = frappe.get_doc({
                "doctype": "Stock Entry",
                "stock_entry_type": "Material Transfer",  # You can set the purpose as per your use case
                "stock_distribution" : sd,
                "company": company,  # Replace with the actual company name
                "to_warehouse" : x,
                "items": update_items
            })
            stock_entry.save(ignore_permissions=True)

        
# Fetch Purchase Order Items
@frappe.whitelist()
def get_purchase_order_items(po_number):
    purchase_order = frappe.get_doc("Purchase Order", po_number)
    items = []
    
    for item in purchase_order.items:
        items.append({
            'item_code': item.item_code,
            'item_name': item.item_name,
            'qty': item.qty,
            'rate': item.rate,
            'warehouse':item.warehouse,
            'uom' : item.uom
        })
    
    return items

@frappe.whitelist()
def get_outlet_items(template):
    temp = frappe.get_doc("Outlet Template", template)
    items = []
    
    for item in temp.outlets:
        items.append({
            'warehouse': item.warehouse,
            'percentage': item.percentage
        })
    
    return items

@frappe.whitelist()
def distribution_excell_generate(doc):
    if not doc: doc = {}
    try:
        doc = json.loads(doc)
    except:
        pass
    data = caluclate_warehouse_percentage_wise_distribution(doc.get("purchase_distribution_items"),doc.get("outlet_selection_table"))
    columns = get_columns(doc)
        
    name = "distribute-"
    if doc.get("purchase_order"):
        name += doc.get("purchase_order")
    else:
        random_word = ''.join(random.choice(string.ascii_lowercase) for _ in range(8))
        name += random_word
    file_name = '{}.xlsx'.format(name)
    generate_excel_and_download(columns, data, file_name, height=20)
    
    return "Ok"
    
def caluclate_warehouse_percentage_wise_distribution(items,warehouse_list,source=None):
    data = []
    for item in items:
        data_dict = {}
        single_data = {}
        total_wr_wise = 0
        if source:
            data_dict["source"] = item.warehouse
        data_dict["item_code"] = item.get("item_code")
        max_percentage = -1  # Assuming all percentages are non-negative
        warehouse_with_max_percentage = None
        
        # loop throw outlet percentage
        for percentage in warehouse_list:
            warehouse = percentage.get("warehouse").lower().replace(" ","_").replace("-","&")
            out_percentage = percentage.get("percentage")
            round_am = round((float(out_percentage) / 100) * float(item.get("qty")))
            single_data[warehouse] = round_am
            total_wr_wise += round_am
            
            # get max percentage warehouse
            if percentage.get("percentage") > max_percentage:
                max_percentage = percentage.get("percentage")
                warehouse_with_max_percentage = percentage.get("warehouse")
        
        
        for key,value in single_data.items():
            data_dict[key] = value
            
        if total_wr_wise != item.get("qty"):
            diff = item.get("qty") - total_wr_wise
            max_percentage_warehouse = warehouse_with_max_percentage.lower().replace(" ","_").replace("-","&")
            max_percentage_warehouse_value = data_dict.get(max_percentage_warehouse) + diff
            data_dict[max_percentage_warehouse] = max_percentage_warehouse_value
        
        data.append(data_dict)
        
    return data
        
def get_columns(doc):
    columns =  [
        {'fieldname': 'item_code', 'label': 'Item Code', 'expwidth': 13, 'width': 90},
    ]
    for warehouse in doc.get("outlet_selection_table"):
        label = warehouse.get("warehouse").lower().replace(" ","_").replace("-","&")
        filedname = warehouse.get("warehouse").lower().replace(" ","_").replace("-","&")
        columns.append({"fieldname": filedname, 'label': label, "expwidth": 13, "width": 90})
        
    return columns
    
    
def generate_excel_and_download(columns, data, file_name, height=25):
    fields, labels, row_widths= [], [], []

    for field in columns:
        if field.get('export', True):
            fields.append(field.get('fieldname'))
            labels.append(field.get('label'))
            row_widths.append(field.get('expwidth', 15))

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    generate_row(worksheet, 1, labels, height=20)
    row_count = 2

    for row in data:
        row_data = [row.get(field) for field in fields]
        generate_row(worksheet, row_count, row_data, widths=row_widths, height=height)
        row_count += 1

    byte_data = io.BytesIO()
    workbook.save(byte_data)

    frappe.local.response.filename = file_name
    frappe.local.response.filecontent = byte_data.getvalue()
    frappe.local.response.type = "download"
    return frappe.local.response


def generate_row(ws, row_count, column_values, font=None, font_size=None, color=None, height=25, widths=None):
    cells = []

    for i, value in enumerate(column_values):
        column_number = i + 1
        cell = ws.cell(row=row_count, column=column_number)
        cell.value = value

        if font:
            cell.font = font
        elif font_size:
            cell.font = Font(size=font_size)

        if color:
            cell.fill = PatternFill(fgColor=color, fill_type='solid')
        if widths:
            ws.column_dimensions[get_column_letter(i + 1)].width = widths[i]

        if isinstance(value, int):
            cell.number_format = "#,##0"
        elif isinstance(value, float):
            cell.number_format = "#,##0.00"

        cell.alignment = Alignment(vertical='center')
        cells.append(cell)

    if height:
        ws.row_dimensions[row_count].height = height

    return cells

def get_site_directory_path():
    site_name = frappe.local.site
    cur_dir = os.getcwd()
    return os.path.join(cur_dir, site_name)
    
@frappe.whitelist()
def get_purchase_receipt(purchase_order):
    receipt = frappe.db.sql("""select pr.name from `tabPurchase Receipt Item` as pri 
                                left join `tabPurchase Receipt` as pr on pri.parent = pr.name 
                                    where pri.purchase_order = '{}' and pr.docstatus = 1 group by pr.name""".format(purchase_order),as_dict=1)
    if receipt and len(receipt) <= 1:
        return receipt[0].get("name")
    elif receipt and len(receipt) > 1:
        receipts = ""
        for x in receipt:
            receipts += x.get("name") + ", "
        frappe.msgprint(str("Receipts {res} Already Created Against Order."
                            .format(res=receipts)),title=_("Multiple Receipt"),indicator="orange")
        return receipt[0].get("name")
    
@frappe.whitelist()
def get_total_from_upload_excell(excell):
    if excell:
        file_path = get_site_directory_path() + excell
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        exclude_column_name = "Item Code"
        total_sum = 0
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if sheet.cell(1, cell.column).value != exclude_column_name:
                    total_sum += cell.value if cell.value is not None else 0

        print("Total Sum (excluding columns with name '{}'): {}".format(exclude_column_name, total_sum))
        workbook.close()
        
        return total_sum